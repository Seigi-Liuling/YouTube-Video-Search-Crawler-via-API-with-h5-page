import re
import logging
import requests
import json
import pandas as pd
from flask import Flask, request, render_template, send_file, redirect, url_for
from googleapiclient.discovery import build
from io import BytesIO
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# 设置日志记录器
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# 配置API_KEY
API_KEY = 'XXXXXXXXXXXXXXXXXXXXXXXXXXX'

# 初始化Flask应用
app = Flask(__name__)

# 搜索关键词视频并处理分页
def search_videos(keyword, max_results=5):
    logging.info(f"搜索关键词: {keyword}, 搜索条数: {max_results}")
    youtube = build('youtube', 'v3', developerKey=API_KEY)
    
    video_ids = []
    next_page_token = None
    while len(video_ids) < max_results:
        remaining_results = min(50, max_results - len(video_ids))
        
        search_response = youtube.search().list(
            q=keyword,
            part='snippet',
            type='video',
            maxResults=remaining_results,
            pageToken=next_page_token
        ).execute()
        
        # 提取视频ID
        video_ids.extend([item['id']['videoId'] for item in search_response['items']])
        
        # 获取下一页的token
        next_page_token = search_response.get('nextPageToken')
        
        # 如果没有更多结果，退出循环
        if not next_page_token:
            break

    logging.debug(f"搜索到的视频ID: {video_ids}")
    return video_ids

# 获取视频详情
def get_video_details(video_ids):
    logging.info(f"获取视频详情，视频ID列表: {video_ids}")
    youtube = build('youtube', 'v3', developerKey=API_KEY)
    
    video_details = []
    
    # 分批处理，每次最多请求50个视频
    for i in range(0, len(video_ids), 50):
        batch_video_ids = video_ids[i:i+50]
        video_response = youtube.videos().list(
            part='snippet,statistics,contentDetails',
            id=','.join(batch_video_ids)
        ).execute()
        
        for item in video_response['items']:
            video_id = item['id']
            video_info = {
                'title': item['snippet']['title'],
                'description': item['snippet']['description'],
                'publishedAt': item['snippet']['publishedAt'],
                'channelTitle': item['snippet']['channelTitle'],
                'tags': ', '.join(item['snippet'].get('tags', [])),
                'view_count': item['statistics'].get('viewCount', 'N/A'),
                'like_count': item['statistics'].get('likeCount', 'N/A'),
                'comment_count': item['statistics'].get('commentCount', 'N/A'),
                'video_id': video_id,
                'duration': item['contentDetails']['duration'],
                'video_link': f'https://www.youtube.com/watch?v={video_id}'  # 新增视频链接
            }
            logging.debug(f"获取到的视频信息: {video_info}")
            video_details.append(video_info)
    
    logging.info(f"共获取到 {len(video_details)} 条视频详情信息")
    return video_details

# 保存为Excel文件，带中文标题和序号列
def save_to_excel(video_details, keyword, num_results):
    logging.info(f"保存视频详情到Excel文件")
    
    # 添加序号列
    for idx, video in enumerate(video_details, start=1):
        video['序号'] = idx
    
    # 将英文字段改为中文
    df = pd.DataFrame(video_details)
    df = df[['序号', 'title', 'video_link', 'description', 'publishedAt', 'channelTitle', 'tags', 'view_count', 'like_count', 'comment_count', 'video_id', 'duration']]
    df.columns = ['序号', '视频标题', '视频链接', '视频描述', '发布时间', '频道名称', '视频tag', '观看量', '点赞量', '评论量', '视频id', '视频时长']
    
    # 格式化文件名
    filename = f"{keyword} {num_results}条 YouTube搜索结果.xlsx"
    
    output = BytesIO()
    
    # 将数据框写入Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # 打开已写入的数据并设置超链接格式
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    # 设置视频链接为超链接并调整样式
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        cell = row[2]  # 第三列是链接列
        cell.hyperlink = cell.value  # 设置超链接
        cell.value = cell.value  # 设置显示文字
        cell.font = Font(color="0000FF", underline="single")  # 蓝色带下划线
    
    # 保存调整后的Excel文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, filename

# 处理用户请求
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        keyword = request.form.get('keyword')
        num_results = int(request.form.get('num_results'))

        if keyword and num_results:
            # 搜索视频
            logging.info(f"收到搜索请求 -> 关键词: {keyword}, 条数: {num_results}")
            video_ids = search_videos(keyword, num_results)
            video_details = get_video_details(video_ids)
            
            # 生成Excel文件
            excel_file, filename = save_to_excel(video_details, keyword, num_results)
            
            # 下载链接
            return send_file(excel_file, download_name=filename, as_attachment=True)
        
    return render_template('index.html')

# 启动应用 设置0.0.0.0监听方便外网访问
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
